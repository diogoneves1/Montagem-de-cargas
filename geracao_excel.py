"""
Geração do arquivo Excel de saída a partir dos containers já montados.
Gera:
  - Aba RESUMO: visão geral de todos os containers (peso, pallets, lotes).
  - Uma aba por container, com o detalhamento linha a linha.
  - Aba PLANO_MONTADO: todos os itens, agrupados por container, com
    subtotal por container — útil para colar/conferir no plano original.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from distribuicao import Container, PESO_MAX, MAX_POSICOES_PISO

AZUL = "0D2C6B"
AZUL_CLARO = "D6E4F7"
CINZA = "F2F2F2"
LARANJA = "FFF2CC"
VERDE = "D9EAD3"
AZUL_MEDIO = "CFE2F3"
AMARELO_TOPO = "FCE4D6"

COR_FAMILIA = {
    "SARDINHA": VERDE,
    "ATUM": AZUL_MEDIO,
    "AZEITE": LARANJA,
}

COR_POSICAO = {
    "BASE": None,             # usa a cor da família
    "BASE (sem topo)": "E2E2E2",  # cinza claro — sinaliza que não sustenta nada
    "TOPO": AMARELO_TOPO,
    "PISO": "FFE5CC",
}


def _fonte(bold=False, tam=10, branco=False):
    return Font(name="Arial", bold=bold, size=tam,
                color="FFFFFF" if branco else "000000")


def _fill(cor_hex):
    return PatternFill("solid", start_color=cor_hex, fgColor=cor_hex)


def _alinhar(h="center", quebra=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=quebra)


def _borda():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _cor_linha(familia: str, posicao: str, lote: str | None) -> str:
    if lote:
        return LARANJA
    cor_pos = COR_POSICAO.get(posicao)
    if cor_pos:
        return cor_pos
    return COR_FAMILIA.get(familia, CINZA)


def _aba_resumo(wb: Workbook, containers: list[Container], titulo: str) -> None:
    ws = wb.active
    ws.title = "RESUMO"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = titulo
    ws["A1"].font = _fonte(True, 14, True)
    ws["A1"].fill = _fill(AZUL)
    ws["A1"].alignment = _alinhar()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = (f"{len(containers)} containers  |  Limite de peso: ~{PESO_MAX:,.0f} kg  |  "
                f"Limite de posições de piso: {MAX_POSICOES_PISO}").replace(",", ".")
    ws["A2"].font = Font(name="Arial", italic=True, size=9)
    ws["A2"].fill = _fill(AZUL_CLARO)
    ws["A2"].alignment = _alinhar()
    ws.row_dimensions[2].height = 16
    ws.append([])

    cabecalhos = ["CNTR", "PESO (kg)", "% PESO", "PALLETS",
                  "POSIÇÕES PISO", "SKUs", "LOTES"]
    ws.append(cabecalhos)
    linha_cab = ws.max_row
    for col in range(1, len(cabecalhos) + 1):
        c = ws.cell(linha_cab, col)
        c.font = _fonte(True, 10, True)
        c.fill = _fill(AZUL)
        c.alignment = _alinhar(quebra=True)
        c.border = _borda()
    ws.row_dimensions[linha_cab].height = 22

    total_peso = total_pallets = 0.0
    for c in containers:
        pct = c.peso / PESO_MAX * 100
        ws.append([
            f"CNTR {c.numero:02d}",
            round(c.peso),
            f"{pct:.1f}%",
            round(c.pallets_total, 2),
            round(c.pPiso, 2),
            c.skus_unicos,
            ", ".join(c.lotes) or "-",
        ])
        total_peso += c.peso
        total_pallets += c.pallets_total
        r = ws.max_row
        bg = CINZA if r % 2 == 0 else "FFFFFF"
        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(r, col)
            cell.font = _fonte()
            cell.fill = _fill(bg)
            cell.border = _borda()
            cell.alignment = _alinhar() if col != 7 else _alinhar("left")
        ws.row_dimensions[r].height = 18

    ws.append(["TOTAL", round(total_peso), "", round(total_pallets, 2), "", "", ""])
    r = ws.max_row
    for col in range(1, len(cabecalhos) + 1):
        cell = ws.cell(r, col)
        cell.font = _fonte(True)
        cell.fill = _fill("D9D9D9")
        cell.alignment = _alinhar()
        cell.border = _borda()
    ws.row_dimensions[r].height = 20

    larguras = [10, 12, 9, 10, 14, 8, 35]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _aba_container(wb: Workbook, c: Container) -> None:
    ws = wb.create_sheet(f"CNTR {c.numero:02d}")
    ws.sheet_view.showGridLines = False

    pct = c.peso / PESO_MAX * 100
    ws.merge_cells("A1:H1")
    ws["A1"] = f"CONTAINER {c.numero:02d}"
    ws["A1"].font = _fonte(True, 13, True)
    ws["A1"].fill = _fill(AZUL)
    ws["A1"].alignment = _alinhar()
    ws.row_dimensions[1].height = 26

    info = (f"Peso: {c.peso:,.0f} kg  ({pct:.1f}%)   |   "
            f"Pallets totais: {c.pallets_total:.2f}   |   "
            f"Posições de piso usadas: {c.pPiso:.2f} / {MAX_POSICOES_PISO}").replace(",", ".")
    ws.merge_cells("A2:H2")
    ws["A2"] = info
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    ws["A2"].fill = _fill(AZUL_CLARO)
    ws["A2"].alignment = _alinhar()
    ws.row_dimensions[2].height = 16
    ws.append([])

    cabecalhos = ["SKU", "PRODUTO", "FAMÍLIA", "POSIÇÃO", "CAIXAS",
                  "PESO (kg)", "PALLETS", "LOTE"]
    ws.append(cabecalhos)
    linha_cab = ws.max_row
    for col in range(1, len(cabecalhos) + 1):
        cell = ws.cell(linha_cab, col)
        cell.font = _fonte(True, 10, True)
        cell.fill = _fill(AZUL)
        cell.alignment = _alinhar(quebra=True)
        cell.border = _borda()
    ws.row_dimensions[linha_cab].height = 20

    linhas_ordenadas = sorted(
        c.linhas, key=lambda l: (l.familia, l.sku, l.posicao)
    )
    for linha in linhas_ordenadas:
        ws.append([
            linha.sku, linha.produto, linha.familia, linha.posicao,
            round(linha.qtd), round(linha.peso), round(linha.pallets, 3),
            linha.lote or "-",
        ])
        r = ws.max_row
        cor = _cor_linha(linha.familia, linha.posicao, linha.lote)
        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(r, col)
            cell.font = _fonte()
            cell.fill = _fill(cor)
            cell.border = _borda()
            cell.alignment = (_alinhar() if col in (1, 4, 5, 6, 7)
                               else _alinhar("left"))
            if col == 6:
                cell.number_format = "#,##0"
        ws.row_dimensions[r].height = 18

    linha_total = ws.max_row + 1
    ws.cell(linha_total, 1).value = "TOTAL"
    ws.cell(linha_total, 5).value = round(sum(l.qtd for l in c.linhas))
    ws.cell(linha_total, 6).value = round(c.peso)
    ws.cell(linha_total, 6).number_format = "#,##0"
    ws.cell(linha_total, 7).value = round(c.pallets_total, 2)
    for col in range(1, len(cabecalhos) + 1):
        cell = ws.cell(linha_total, col)
        cell.font = _fonte(True)
        cell.fill = _fill("D9D9D9")
        cell.alignment = _alinhar()
        cell.border = _borda()
    ws.row_dimensions[linha_total].height = 18

    legenda_linha = linha_total + 2
    ws.cell(legenda_linha, 1).value = "LEGENDA:"
    ws.cell(legenda_linha, 1).font = _fonte(True)
    legendas = [
        (LARANJA, "SKU com lote específico"),
        (AMARELO_TOPO, "TOPO — empilhado sobre BASE"),
        ("FFE5CC", "PISO — posição própria (frágil sem base)"),
        ("E2E2E2", "BASE (sem topo) — ex: AZEITE, pallet não sustenta nada em cima"),
        (VERDE, "Família SARDINHA"),
        (AZUL_MEDIO, "Família ATUM"),
    ]
    for i, (cor, texto) in enumerate(legendas):
        ws.cell(legenda_linha + 1 + i, 1).fill = _fill(cor)
        ws.cell(legenda_linha + 1 + i, 1).value = texto
        ws.cell(legenda_linha + 1 + i, 1).font = _fonte()

    larguras = [9, 58, 12, 22, 9, 13, 10, 10]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _aba_plano_montado(wb: Workbook, containers: list[Container]) -> None:
    """Aba com todos os itens, agrupados por container e com subtotal —
    pronta para copiar/colar de volta no plano original, preservando
    a informação de LOTE em destaque para a montagem de pedidos."""
    ws = wb.create_sheet("PLANO_MONTADO")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"] = "PLANO MONTADO — TODOS OS CONTAINERS"
    ws["A1"].font = _fonte(True, 13, True)
    ws["A1"].fill = _fill(AZUL)
    ws["A1"].alignment = _alinhar()
    ws.row_dimensions[1].height = 26
    ws.append([])

    cabecalhos = ["CONTAINER", "SKU", "PRODUTO", "FAMÍLIA", "POSIÇÃO",
                  "CAIXAS", "PESO (kg)", "PALLETS", "LOTE"]
    ws.append(cabecalhos)
    linha_cab = ws.max_row
    for col in range(1, len(cabecalhos) + 1):
        cell = ws.cell(linha_cab, col)
        cell.font = _fonte(True, 10, True)
        cell.fill = _fill(AZUL)
        cell.alignment = _alinhar(quebra=True)
        cell.border = _borda()
    ws.row_dimensions[linha_cab].height = 20

    for c in containers:
        linhas_ordenadas = sorted(c.linhas, key=lambda l: (l.familia, l.sku))
        for linha in linhas_ordenadas:
            ws.append([
                f"CNTR {c.numero:02d}", linha.sku, linha.produto, linha.familia,
                linha.posicao, round(linha.qtd), round(linha.peso),
                round(linha.pallets, 3), linha.lote or "-",
            ])
            r = ws.max_row
            cor = _cor_linha(linha.familia, linha.posicao, linha.lote)
            for col in range(1, len(cabecalhos) + 1):
                cell = ws.cell(r, col)
                cell.font = _fonte()
                cell.fill = _fill(cor)
                cell.border = _borda()
                cell.alignment = (_alinhar() if col in (1, 2, 5, 6, 7, 8)
                                   else _alinhar("left"))
                if col == 7:
                    cell.number_format = "#,##0"
            ws.row_dimensions[r].height = 17

        # Subtotal do container
        r = ws.max_row + 1
        ws.merge_cells(f"A{r}:D{r}")
        ws.cell(r, 1).value = f"TOTAL CNTR {c.numero:02d}"
        ws.cell(r, 6).value = round(sum(l.qtd for l in c.linhas))
        ws.cell(r, 7).value = round(c.peso)
        ws.cell(r, 7).number_format = "#,##0"
        ws.cell(r, 8).value = round(c.pallets_total, 2)
        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(r, col)
            cell.font = _fonte(True, branco=True)
            cell.fill = _fill("2E75B6")
            cell.alignment = _alinhar()
            cell.border = _borda()
        ws.row_dimensions[r].height = 18
        ws.append([])

    larguras = [11, 9, 58, 12, 14, 9, 13, 10, 10]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar_excel(containers: list[Container], titulo: str = "PLANO DE EMBARQUE") -> bytes:
    """Gera o arquivo .xlsx completo (RESUMO + abas por container +
    PLANO_MONTADO) e retorna os bytes prontos para download."""
    wb = Workbook()
    _aba_resumo(wb, containers, titulo)
    for c in containers:
        _aba_container(wb, c)
    _aba_plano_montado(wb, containers)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
